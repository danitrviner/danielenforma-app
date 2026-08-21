#!/usr/bin/env python3
"""
Normaliza el Excel de 840 ejercicios (tarea: reemplazo del catálogo, agosto 2026)
a scripts/out/ejercicios-catalogo.json, con la forma exacta de Omit<Exercise, 'id'>.

- Separa "Nombre - Grupo1, Grupo2, Grupo3" en nombre + lista de grupos.
- Corrige a mano las 4 filas que no siguen ese patrón (variantes de Spoto press).
- Canonicaliza las ~45 etiquetas de grupo en bruto (mayúsculas/sinónimos) al
  MuscleGroup tipado de src/types.ts — incluye las 2 claves nuevas (lumbares,
  rotadores). El primer grupo listado es el principal (muscleGroup); el resto
  va a secondaryMuscleGroups.
- Los "Fullbody" (ejercicios olímpicos) quedan sin muscleGroup: no hay reparto
  posible con este sistema.

No escribe nada en Firestore — eso lo hace importarEjerciciosCatalogo.mjs, que
lee el JSON que este script genera.

Uso:
    python3 scripts/normalizarEjerciciosExcel.py [ruta_al_xlsx]
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

DEFAULT_XLSX = "/Users/dani/Downloads/Hoja de cálculo sin título (3).xlsx"
OUT_PATH = Path(__file__).parent / "out" / "ejercicios-catalogo.json"

# Filas rotas en el Excel (no siguen "Nombre - Grupos"): nombre correcto -> grupos correctos.
MANUAL_FIXES = {
    "Spoto press (caja torácica amplia) Pectoral, deltoides anterior, tríceps":
        ("Spoto press (caja torácica amplia)", "Pectoral, deltoides anterior, tríceps"),
    "Spoto press (caja torácica plana) Pectoral, deltoides anterior, tríceps":
        ("Spoto press (caja torácica plana)", "Pectoral, deltoides anterior, tríceps"),
    "Spoto press + goma (10-20 kg) (caja torácica amplia)- Pectoral, deltoides anterior, tríceps":
        ("Spoto press + goma (10-20 kg) (caja torácica amplia)", "Pectoral, deltoides anterior, tríceps"),
    "Spoto press + goma (10-20 kg) (caja torácica plana)- Pectoral, deltoides anterior, tríceps":
        ("Spoto press + goma (10-20 kg) (caja torácica plana)", "Pectoral, deltoides anterior, tríceps"),
    "Spoto press en multipower - (caja torácica amplia) Pectoral, deltoides anterior, tríceps":
        ("Spoto press en multipower (caja torácica amplia)", "Pectoral, deltoides anterior, tríceps"),
    "Spoto press en multipower - (caja torácica plana) Pectoral, deltoides anterior, tríceps":
        ("Spoto press en multipower (caja torácica plana)", "Pectoral, deltoides anterior, tríceps"),
}

# Canonicalización case-insensitive -> MuscleGroup. 'fullbody' se mapea a None
# a propósito (sin grupo, no computable).
GROUP_MAP = {
    "pecho": "pecho", "pectoral": "pecho",
    "dorsal": "dorsal",
    "trapecio": "trapecio", "complejo escapular": "trapecio",
    "deltoides anterior": "deltoide_ant",
    "deltoides medial": "deltoide_lat",
    "deltoides posterior": "deltoide_post",
    "biceps": "biceps",
    "triceps": "triceps",
    "antebrazo": "antebrazo",
    "cuadriceps": "cuadriceps",
    "isquios": "isquios", "isquiotibiales": "isquios",
    "gluteo": "gluteo", "gluteo (gluteo medio)": "gluteo",
    "aductores": "aductores",
    "gemelo": "gemelo", "gemelos": "gemelo", "soleos": "gemelo",
    "abdomen": "core", "oblicuos": "core", "psoas iliaco": "core",
    "lumbares": "lumbares", "erectores espinales": "lumbares",
    "rotadores externos": "rotadores", "rotadores internos": "rotadores",
    "manguito rotador": "rotadores",
    "fullbody": None,
}


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def canonical_key(raw_group: str) -> str:
    """Normaliza una etiqueta en bruto a la clave de búsqueda en GROUP_MAP."""
    return strip_accents(raw_group.strip().lower())


def slugify(name: str) -> str:
    s = strip_accents(name).lower()
    s = s.replace("&", " y ")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def parse_row(name_and_groups: str):
    if name_and_groups in MANUAL_FIXES:
        name, groups_str = MANUAL_FIXES[name_and_groups]
    elif " - " in name_and_groups:
        name, groups_str = name_and_groups.rsplit(" - ", 1)
    else:
        return None, None, f"sin separador ' - ': {name_and_groups!r}"

    raw_groups = [g.strip() for g in groups_str.split(",")]
    mapped = []
    unmapped = []
    for g in raw_groups:
        key = canonical_key(g)
        if key not in GROUP_MAP:
            unmapped.append(g)
            continue
        target = GROUP_MAP[key]
        if target is not None and target not in mapped:
            mapped.append(target)
    return name.strip(), mapped, ("; ".join(f"grupo sin mapeo: {g!r}" for g in unmapped) if unmapped else None)


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    exercises = []
    issues = []
    seen_ids = {}
    no_group_count = 0

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=4, values_only=True):
        name_and_groups, video_url, coach_video = row
        if not name_and_groups:
            continue

        name, groups, problem = parse_row(name_and_groups)
        if problem:
            issues.append(f"{name_and_groups!r} -> {problem}")
            if not groups:
                continue

        if not video_url:
            issues.append(f"{name!r}: sin URL de vídeo")

        ex_id = f"sys_{slugify(name)}"
        if ex_id in seen_ids:
            # Dos filas con el mismo nombre base pero grupos distintos (p.ej.
            # "Rotación de hombro a 90º" externa vs interna) — el nombre solo
            # se distingue por el grupo, así que se lo añadimos entre paréntesis
            # en vez de un sufijo numérico sin sentido para el coach.
            original_groups = name_and_groups.rsplit(" - ", 1)[-1] if " - " in name_and_groups else ""
            disambiguated = f"{name} ({original_groups.strip()})" if original_groups else name
            issues.append(f"{name!r}: nombre duplicado ({ex_id}) — renombrado a {disambiguated!r}")
            name = disambiguated
            ex_id = f"sys_{slugify(name)}"
            n = 2
            while ex_id in seen_ids:
                ex_id = f"sys_{slugify(name)}-{n}"
                n += 1
        seen_ids[ex_id] = name

        muscle_group = groups[0] if groups else None
        secondary = groups[1:] if len(groups) > 1 else []
        if muscle_group is None:
            no_group_count += 1

        exercise = {
            "id": ex_id,
            "ownerId": "system",
            "name": name,
            "primaryFocus": name,
            "type": "fuerza",
            "isCustom": False,
            "videoUrl": video_url or None,
        }
        if muscle_group:
            exercise["muscleGroup"] = muscle_group
        if secondary:
            exercise["secondaryMuscleGroups"] = secondary

        exercises.append(exercise)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(exercises, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Total ejercicios generados: {len(exercises)}")
    print(f"Sin grupo muscular (fullbody u otros): {no_group_count}")
    print(f"Escrito en: {OUT_PATH}")
    if issues:
        print(f"\nIncidencias ({len(issues)}):")
        for i in issues:
            print(f"  - {i}")
    else:
        print("\nSin incidencias.")


if __name__ == "__main__":
    main()
